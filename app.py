from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
from io import BytesIO
import os
import datetime
import anthropic
import io
from reportlab.pdfgen import canvas
from reportlab.lib.colors import black
from pypdf import PdfReader, PdfWriter, Transformation
from pypdf.generic import RectangleObject

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, 'excel_templates')

DOCTOR = {
    'nome': 'FÁBIO BATISTELLA',
    'crm': '31746',
    'especialidade': 'REUMATOLOGIA',
    'telefone': '54 996978356',
}

FARMACOS_BIOLOGICOS = [
    'Adalimumabe 40mg',
    'Adalimumabe biossimilar 40mg',
    'Certolizumabe pegol 200mg',
    'Etanercepte 50mg',
    'Golimumabe 50mg SC',
    'Infliximabe 100mg/10ml IV',
    'Tocilizumabe 20mg/ml IV',
    'Tocilizumabe 162mg SC',
    'Secuquinumabe 150mg SC',
    'Abatacepte 125mg SC',
    'Abatacepte 250mg IV',
    'Baricitinibe 2mg',
    'Baricitinibe 4mg',
    'Tofacitinibe 5mg',
    'Upadacitinibe 15mg',
    'Rituximabe',
    'Ixequizumabe 80mg SC',
    'Guselcumabe 100mg SC',
]

# Fármacos listados nos termos de consentimento originais
FARMACOS_TC_AR = [
    'Adalimumabe', 'Certolizumabe pegol', 'Etanercepte',
    'Golimumabe', 'Infliximabe', 'Abatacepte', 'Tocilizumabe', 'Rituximabe',
]
FARMACOS_TC_AP = [
    'Adalimumabe', 'Certolizumabe pegol', 'Etanercepte',
    'Golimumabe', 'Infliximabe', 'Abatacepte', 'Secuquinumabe',
]
FARMACOS_TC_EA = [
    'Adalimumabe', 'Certolizumabe pegol', 'Etanercepte',
    'Golimumabe', 'Infliximabe', 'Secuquinumabe', 'Ixequizumabe',
]

FARMACOS_SINTETICOS = [
    'Metotrexato',
    'Hidroxicloroquina',
    'Sulfassalazina',
    'Leflunomida',
    'Azatioprina',
]

CID_AR = ['M05.0', 'M05.1', 'M05.2', 'M05.3', 'M05.8', 'M05.9', 'M06.0', 'M06.1', 'M06.2', 'M06.8', 'M06.9', 'M08.0']
CID_AP = ['M07.0', 'M07.1', 'M07.2', 'M07.3', 'M073']
CID_EA = ['M45', 'M45.0', 'M46.0', 'M46.1', 'M46.2', 'M46.8', 'M46.9']

FARMACOS_BIOLOGICOS_EA = [
    'Adalimumabe 40mg',
    'Adalimumabe biossimilar 40mg',
    'Certolizumabe pegol 200mg',
    'Etanercepte 50mg',
    'Golimumabe 50mg SC',
    'Infliximabe 100mg/10ml IV',
    'Secuquinumabe 150mg SC',
    'Ixequizumabe 80mg SC',
]


def wb_write(ws, coord, value):
    if value is not None and str(value).strip() != '':
        try:
            ws[coord] = value
        except Exception:
            pass


def sim_nao(value, prefix=True):
    if value == 'Sim':
        return '1. Sim' if prefix else 'Sim'
    return '2. Não' if prefix else 'Não'


def generate_ar_inicial(data):
    wb = openpyxl.load_workbook(os.path.join(EXCEL_DIR, 'ar_inicial.xlsx'))
    ws = wb.active

    wb_write(ws, 'C4', DOCTOR['nome'])
    wb_write(ws, 'C5', DOCTOR['crm'])
    wb_write(ws, 'E5', DOCTOR['especialidade'])
    wb_write(ws, 'C6', DOCTOR['telefone'])

    wb_write(ws, 'C8', data.get('nome_paciente', ''))
    wb_write(ws, 'C9', data.get('idade', ''))
    wb_write(ws, 'E9', data.get('sexo', ''))
    wb_write(ws, 'C10', data.get('telefone_paciente', ''))

    wb_write(ws, 'C12', data.get('data_diagnostico', ''))
    wb_write(ws, 'C13', data.get('cid10', ''))

    fr_res = data.get('fr_resultado', '')
    wb_write(ws, 'C15', '1. Positivo' if fr_res == 'Positivo' else '2. Negativo')
    wb_write(ws, 'E15', data.get('fr_valor', ''))

    anticcp_res = data.get('anticcp_resultado', '')
    wb_write(ws, 'C16', '1. Positivo' if anticcp_res == 'Positivo' else '2. Negativo')
    wb_write(ws, 'E16', data.get('anticcp_valor', ''))

    rx_maos = data.get('rx_maos_realizado', 'Não')
    wb_write(ws, 'E17', sim_nao(rx_maos))
    if rx_maos == 'Sim':
        wb_write(ws, 'C18', sim_nao(data.get('rx_maos_erosoes', 'Não')))
        wb_write(ws, 'E18', sim_nao(data.get('rx_maos_diminuicao', 'Não')))

    rmn = data.get('rmn_realizado', 'Não')
    wb_write(ws, 'E19', sim_nao(rmn))
    if rmn == 'Sim':
        wb_write(ws, 'C20', sim_nao(data.get('rmn_erosoes', 'Não')))
        wb_write(ws, 'E20', sim_nao(data.get('rmn_diminuicao', 'Não')))
        wb_write(ws, 'C21', sim_nao(data.get('rmn_sinovite', 'Não')))

    wb_write(ws, 'C24', data.get('indice_tipo', ''))
    wb_write(ws, 'D26', data.get('articulacoes_dor', ''))
    wb_write(ws, 'D27', data.get('articulacoes_edema', ''))
    wb_write(ws, 'D28', data.get('eva_paciente', ''))
    wb_write(ws, 'D29', data.get('eva_medico', ''))
    wb_write(ws, 'D30', data.get('pcr', ''))
    wb_write(ws, 'D31', data.get('vsg', ''))
    wb_write(ws, 'D32', data.get('valor_indice', ''))

    # ACR-EULAR 2010 - Envolvimento articular
    joint_map = {
        '1_grande': ('C35', 0),
        '2_10_grandes': ('C36', 1),
        '1_3_pequenas': ('C37', 2),
        '4_10_pequenas': ('C38', 3),
        '10_mais': ('C39', 5),
    }
    joint = data.get('acr_articulacoes', '')
    if joint in joint_map:
        cell, val = joint_map[joint]
        wb_write(ws, cell, val)

    # ACR-EULAR 2010 - Sorologia
    sero_map = {
        'negativos': ('C40', 0),
        'baixos': ('C41', 2),
        'altos': ('C42', 3),
    }
    sero = data.get('acr_sorologia', '')
    if sero in sero_map:
        cell, val = sero_map[sero]
        wb_write(ws, cell, val)

    # ACR-EULAR 2010 - Fase aguda
    fase = data.get('acr_fase_aguda', '')
    if fase == 'normais':
        wb_write(ws, 'C43', 0)
    elif fase == 'alterados':
        wb_write(ws, 'C44', 1)

    # ACR-EULAR 2010 - Duração
    duracao = data.get('acr_duracao', '')
    if duracao == 'menos_6':
        wb_write(ws, 'C45', 0)
    elif duracao == 'mais_6':
        wb_write(ws, 'C46', 1)

    # Tratamento anterior (até 4 linhas)
    for i in range(1, 5):
        row = 49 + i
        farmaco = data.get(f'farmaco_ant_{i}', '')
        if farmaco:
            wb_write(ws, f'B{row}', farmaco)
            wb_write(ws, f'C{row}', data.get(f'posologia_ant_{i}', ''))
            wb_write(ws, f'E{row}', data.get(f'periodo_ant_{i}', ''))

    wb_write(ws, 'C55', data.get('farmaco_proposto', ''))
    wb_write(ws, 'C56', data.get('posologia_proposta', ''))
    wb_write(ws, 'C57', data.get('peso', ''))
    wb_write(ws, 'C58', sim_nao(data.get('ppd_rx', 'Não')))
    wb_write(ws, 'B60', data.get('observacoes', ''))
    wb_write(ws, 'C64', data.get('data', datetime.date.today().strftime('%d.%m.%Y')))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_ar_manutencao(data):
    wb = openpyxl.load_workbook(os.path.join(EXCEL_DIR, 'ar_manutencao.xlsx'))
    ws = wb.active

    wb_write(ws, 'C4', DOCTOR['nome'])
    wb_write(ws, 'C5', DOCTOR['crm'])
    wb_write(ws, 'E5', DOCTOR['especialidade'])
    wb_write(ws, 'C6', DOCTOR['telefone'])

    wb_write(ws, 'C8', data.get('nome_paciente', ''))
    wb_write(ws, 'C9', data.get('idade', ''))
    wb_write(ws, 'E9', data.get('sexo', ''))
    wb_write(ws, 'C10', data.get('inicio_tratamento', ''))
    wb_write(ws, 'C11', data.get('cid10', ''))
    wb_write(ws, 'C12', data.get('peso', ''))

    wb_write(ws, 'D14', data.get('indice_tipo', ''))
    wb_write(ws, 'D16', data.get('articulacoes_dor', ''))
    wb_write(ws, 'D17', data.get('articulacoes_edema', ''))
    wb_write(ws, 'D18', data.get('eva_paciente', ''))
    wb_write(ws, 'D19', data.get('eva_medico', ''))
    wb_write(ws, 'D20', data.get('pcr', ''))
    wb_write(ws, 'D21', data.get('vsg', ''))
    wb_write(ws, 'D22', data.get('valor_indice', ''))

    wb_write(ws, 'C24', sim_nao(data.get('boa_resposta', 'Não')))
    wb_write(ws, 'B26', data.get('descricao_falha', ''))

    wb_write(ws, 'D28', sim_nao(data.get('manter', 'Não')))
    wb_write(ws, 'D29', sim_nao(data.get('modificar', 'Não')))
    wb_write(ws, 'D30', data.get('farmaco', ''))
    wb_write(ws, 'D31', data.get('posologia', ''))

    sintetico = data.get('sintetico_associado', 'Não')
    wb_write(ws, 'D32', sim_nao(sintetico))
    if sintetico == 'Sim':
        wb_write(ws, 'D33', data.get('qual_sintetico', ''))
        wb_write(ws, 'D34', data.get('posologia_sintetico', ''))

    wb_write(ws, 'B36', data.get('observacoes', ''))
    wb_write(ws, 'C38', data.get('data', datetime.date.today().strftime('%d.%m.%Y')))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_ap_inicial(data):
    wb = openpyxl.load_workbook(os.path.join(EXCEL_DIR, 'ap_inicial.xlsx'))
    ws = wb.active

    wb_write(ws, 'C4', DOCTOR['nome'])
    wb_write(ws, 'C5', DOCTOR['crm'])
    wb_write(ws, 'E5', DOCTOR['especialidade'])
    wb_write(ws, 'C6', DOCTOR['telefone'])

    wb_write(ws, 'C8', data.get('nome_paciente', ''))
    wb_write(ws, 'C9', data.get('idade', ''))
    wb_write(ws, 'E9', data.get('sexo', ''))
    wb_write(ws, 'C10', data.get('telefone_paciente', ''))

    wb_write(ws, 'C12', data.get('data_diagnostico', ''))
    wb_write(ws, 'C13', data.get('cid10', ''))
    wb_write(ws, 'E15', data.get('fr_valor', ''))
    wb_write(ws, 'D16', sim_nao(data.get('formacao_ossea', 'Não'), prefix=False))

    rx_maos = data.get('rx_maos_realizado', 'Não')
    wb_write(ws, 'E17', sim_nao(rx_maos))
    if rx_maos == 'Sim':
        wb_write(ws, 'C18', sim_nao(data.get('rx_maos_erosoes', 'Não')))
        wb_write(ws, 'E18', sim_nao(data.get('rx_maos_diminuicao', 'Não')))

    rx_axial = data.get('rx_axial_realizado', 'Não')
    wb_write(ws, 'E19', sim_nao(rx_axial))
    if rx_axial == 'Sim':
        wb_write(ws, 'C20', sim_nao(data.get('sacroileite', 'Não')))
        wb_write(ws, 'E20', sim_nao(data.get('sindesmofitos', 'Não')))

    rmn_eco = data.get('rmn_eco_realizado', 'Não')
    wb_write(ws, 'E21', sim_nao(rmn_eco))
    if rmn_eco == 'Sim':
        wb_write(ws, 'C22', sim_nao(data.get('sinovite', 'Não')))
        wb_write(ws, 'E22', sim_nao(data.get('tenossinovite', 'Não')))
        wb_write(ws, 'C23', sim_nao(data.get('entesopatia', 'Não')))

    rmn_sacro = data.get('rmn_sacro_realizado', 'Não')
    wb_write(ws, 'E24', sim_nao(rmn_sacro))
    if rmn_sacro == 'Sim':
        wb_write(ws, 'C25', sim_nao(data.get('edema_osseo', 'Não')))

    wb_write(ws, 'D27', data.get('indice_tipo', ''))
    wb_write(ws, 'D28', data.get('pcr', ''))
    wb_write(ws, 'D29', data.get('vsg', ''))
    wb_write(ws, 'D30', data.get('valor_indice', ''))

    # CASPAR
    psoriase_atual = data.get('psoriase_atual') == 'Sim'
    hist_pessoal = data.get('hist_pessoal') == 'Sim'
    hist_familiar = data.get('hist_familiar') == 'Sim'
    distrofia = data.get('distrofia_ungueal') == 'Sim'
    fr_neg = data.get('fr_negativo') == 'Sim'
    dactilite = data.get('dactilite') == 'Sim'
    formacao_ossea_rx = data.get('formacao_ossea_rx') == 'Sim'

    wb_write(ws, 'C32', 'Sim' if psoriase_atual else 'Não')
    wb_write(ws, 'D32', 2 if psoriase_atual else 0)

    wb_write(ws, 'C33', 'Sim' if hist_pessoal else 'Não')
    wb_write(ws, 'D33', 1 if hist_pessoal and not psoriase_atual else 0)

    wb_write(ws, 'C34', 'Sim' if hist_familiar else 'Não')
    wb_write(ws, 'D34', 1 if hist_familiar else 0)

    wb_write(ws, 'C35', 'Sim' if distrofia else 'Não')
    wb_write(ws, 'D35', 1 if distrofia else 0)

    wb_write(ws, 'C36', 'Negativo' if fr_neg else 'Positivo')
    wb_write(ws, 'D36', 1 if fr_neg else 0)

    wb_write(ws, 'C37', 'Sim' if dactilite else 'Não')
    wb_write(ws, 'D37', 1 if dactilite else 0)

    wb_write(ws, 'C38', 'Sim' if formacao_ossea_rx else 'Não')
    wb_write(ws, 'D38', 1 if formacao_ossea_rx else 0)

    for i in range(1, 4):
        row = 41 + i
        farmaco = data.get(f'farmaco_ant_{i}', '')
        if farmaco:
            wb_write(ws, f'B{row}', farmaco)
            wb_write(ws, f'C{row}', data.get(f'posologia_ant_{i}', ''))
            wb_write(ws, f'E{row}', data.get(f'periodo_ant_{i}', ''))

    wb_write(ws, 'C47', data.get('farmaco_proposto', ''))
    wb_write(ws, 'C48', data.get('posologia_proposta', ''))
    wb_write(ws, 'C49', data.get('peso', ''))
    wb_write(ws, 'C50', sim_nao(data.get('ppd_rx', 'Não')))
    wb_write(ws, 'B52', data.get('observacoes', ''))
    wb_write(ws, 'C56', data.get('data', datetime.date.today().strftime('%d.%m.%Y')))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_ea(data):
    tipo = data.get('tipo', 'Inicial')
    if tipo == 'Manutencao':
        wb = openpyxl.load_workbook(os.path.join(EXCEL_DIR, 'ea_manutencao.xlsx'))
    else:
        wb = openpyxl.load_workbook(os.path.join(EXCEL_DIR, 'ea_inicial.xlsx'))
    ws = wb.active

    wb_write(ws, 'C4', DOCTOR['nome'])
    wb_write(ws, 'C5', DOCTOR['crm'])
    wb_write(ws, 'E5', DOCTOR['especialidade'])
    wb_write(ws, 'C6', DOCTOR['telefone'])

    wb_write(ws, 'C8', data.get('nome_paciente', ''))
    wb_write(ws, 'C9', data.get('idade', ''))
    wb_write(ws, 'E9', data.get('sexo', ''))
    wb_write(ws, 'C10', data.get('telefone_paciente', '') if tipo == 'Inicial' else data.get('inicio_tratamento', ''))
    wb_write(ws, 'C11', data.get('cid10', ''))
    wb_write(ws, 'C12', data.get('peso', ''))

    if tipo == 'Inicial':
        wb_write(ws, 'C14', data.get('data_diagnostico', ''))
        wb_write(ws, 'C15', data.get('hlab27_resultado', ''))
        wb_write(ws, 'C16', data.get('forma_doenca', ''))
        wb_write(ws, 'C17', data.get('criterio_diagnostico', ''))

        rx_sacro = data.get('rx_sacro_realizado', 'Não')
        wb_write(ws, 'E20', sim_nao(rx_sacro))
        if rx_sacro == 'Sim':
            wb_write(ws, 'C21', data.get('sacroileite_grau', ''))

        rmn_sacro = data.get('rmn_sacro_realizado', 'Não')
        wb_write(ws, 'E22', sim_nao(rmn_sacro))
        if rmn_sacro == 'Sim':
            wb_write(ws, 'C23', sim_nao(data.get('rmn_edema_osseo', 'Não')))
            wb_write(ws, 'E23', sim_nao(data.get('rmn_erosoes', 'Não')))
            wb_write(ws, 'C24', sim_nao(data.get('rmn_esclerose', 'Não')))

        rx_coluna = data.get('rx_coluna_realizado', 'Não')
        wb_write(ws, 'E25', sim_nao(rx_coluna))
        if rx_coluna == 'Sim':
            wb_write(ws, 'C26', sim_nao(data.get('sindesmofitos', 'Não')))
            wb_write(ws, 'E26', sim_nao(data.get('coluna_bambu', 'Não')))

        wb_write(ws, 'C29', data.get('pcr', ''))
        wb_write(ws, 'C30', data.get('vsg', ''))
        wb_write(ws, 'C31', data.get('basdai_q1', ''))
        wb_write(ws, 'C32', data.get('basdai_q2', ''))
        wb_write(ws, 'C33', data.get('basdai_q3', ''))
        wb_write(ws, 'C34', data.get('basdai_q4', ''))
        wb_write(ws, 'C35', data.get('basdai_q5', ''))
        wb_write(ws, 'C36', data.get('basdai_q6', ''))
        wb_write(ws, 'C37', data.get('asdas_pga', ''))
        wb_write(ws, 'C38', data.get('indice_tipo', ''))
        wb_write(ws, 'C39', data.get('valor_indice', ''))

        for i in range(1, 5):
            row = 42 + i
            farmaco = data.get(f'farmaco_ant_{i}', '')
            if farmaco:
                wb_write(ws, f'B{row}', farmaco)
                wb_write(ws, f'C{row}', data.get(f'posologia_ant_{i}', ''))
                wb_write(ws, f'E{row}', data.get(f'periodo_ant_{i}', ''))

        wb_write(ws, 'C49', data.get('farmaco_proposto', ''))
        wb_write(ws, 'C50', data.get('posologia_proposta', ''))
        wb_write(ws, 'C51', sim_nao(data.get('ppd_rx', 'Não')))
        wb_write(ws, 'B53', data.get('observacoes', ''))
        wb_write(ws, 'C57', data.get('data', datetime.date.today().strftime('%d.%m.%Y')))

    else:  # Manutenção
        wb_write(ws, 'C14', data.get('indice_tipo', ''))
        wb_write(ws, 'C15', data.get('pcr', ''))
        wb_write(ws, 'C16', data.get('vsg', ''))
        wb_write(ws, 'C17', data.get('basdai_q1', ''))
        wb_write(ws, 'C18', data.get('basdai_q2', ''))
        wb_write(ws, 'C19', data.get('basdai_q3', ''))
        wb_write(ws, 'C20', data.get('basdai_q4', ''))
        wb_write(ws, 'C21', data.get('basdai_q5', ''))
        wb_write(ws, 'C22', data.get('basdai_q6', ''))
        wb_write(ws, 'C23', data.get('asdas_pga', ''))
        wb_write(ws, 'C24', data.get('valor_indice', ''))

        wb_write(ws, 'C26', sim_nao(data.get('boa_resposta', 'Não')))
        wb_write(ws, 'B28', data.get('descricao_falha', ''))

        wb_write(ws, 'D30', sim_nao(data.get('manter', 'Não')))
        wb_write(ws, 'D31', sim_nao(data.get('modificar', 'Não')))
        wb_write(ws, 'D32', data.get('farmaco', ''))
        wb_write(ws, 'D33', data.get('posologia', ''))

        wb_write(ws, 'B35', data.get('observacoes', ''))
        wb_write(ws, 'C37', data.get('data', datetime.date.today().strftime('%d.%m.%Y')))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_ap_manutencao(data):
    wb = openpyxl.load_workbook(os.path.join(EXCEL_DIR, 'ap_manutencao.xlsx'))
    ws = wb.active

    wb_write(ws, 'C4', DOCTOR['nome'])
    wb_write(ws, 'C5', DOCTOR['crm'])
    wb_write(ws, 'E5', DOCTOR['especialidade'])
    wb_write(ws, 'C6', DOCTOR['telefone'])

    wb_write(ws, 'C8', data.get('nome_paciente', ''))
    wb_write(ws, 'C9', data.get('idade', ''))
    wb_write(ws, 'E9', data.get('sexo', ''))
    wb_write(ws, 'C10', data.get('inicio_tratamento', ''))
    wb_write(ws, 'C11', data.get('cid10', ''))
    wb_write(ws, 'C12', data.get('peso', ''))

    wb_write(ws, 'D14', data.get('indice_tipo', ''))
    wb_write(ws, 'D16', data.get('pcr', ''))
    wb_write(ws, 'D17', data.get('vsg', ''))
    wb_write(ws, 'D18', data.get('valor_indice', ''))

    wb_write(ws, 'D20', sim_nao(data.get('boa_resposta', 'Não'), prefix=False))
    prm = data.get('prm', 'Não')
    wb_write(ws, 'D21', sim_nao(prm, prefix=False))
    wb_write(ws, 'B23', data.get('descricao_prm', ''))

    wb_write(ws, 'E25', sim_nao(data.get('manter', 'Não'), prefix=False))
    wb_write(ws, 'E26', sim_nao(data.get('modificar', 'Não'), prefix=False))
    wb_write(ws, 'E27', data.get('farmaco', ''))
    wb_write(ws, 'E28', data.get('posologia', ''))

    sintetico = data.get('sintetico_associado', 'Não')
    wb_write(ws, 'E29', sim_nao(sintetico, prefix=False))
    if sintetico == 'Sim':
        wb_write(ws, 'E30', data.get('qual_sintetico', ''))
        wb_write(ws, 'E31', data.get('posologia_sintetico', ''))

    wb_write(ws, 'B33', data.get('observacoes', ''))
    wb_write(ws, 'C35', data.get('data', datetime.date.today().strftime('%d.%m.%Y')))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_filename(prefix, nome, data):
    nome_clean = nome.replace(' ', '_').upper() if nome else 'PACIENTE'
    return f'{prefix}_{nome_clean}_{data}.xlsx'


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/ar/inicial', methods=['GET', 'POST'])
def ar_inicial():
    if request.method == 'POST':
        data = request.form.to_dict()
        buf = generate_ar_inicial(data)
        nome = data.get('nome_paciente', 'paciente')
        today = datetime.date.today().strftime('%Y%m%d')
        filename = make_filename('AR_Inicial', nome, today)
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('ar_inicial.html',
                           farmacos=FARMACOS_BIOLOGICOS,
                           sinteticos=FARMACOS_SINTETICOS,
                           cids=CID_AR,
                           today=datetime.date.today().strftime('%d.%m.%Y'))


@app.route('/ar/manutencao', methods=['GET', 'POST'])
def ar_manutencao():
    if request.method == 'POST':
        data = request.form.to_dict()
        buf = generate_ar_manutencao(data)
        nome = data.get('nome_paciente', 'paciente')
        today = datetime.date.today().strftime('%Y%m%d')
        filename = make_filename('AR_Manutencao', nome, today)
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('ar_manutencao.html',
                           farmacos=FARMACOS_BIOLOGICOS,
                           sinteticos=FARMACOS_SINTETICOS,
                           cids=CID_AR,
                           today=datetime.date.today().strftime('%d.%m.%Y'))


@app.route('/ap/inicial', methods=['GET', 'POST'])
def ap_inicial():
    if request.method == 'POST':
        data = request.form.to_dict()
        buf = generate_ap_inicial(data)
        nome = data.get('nome_paciente', 'paciente')
        today = datetime.date.today().strftime('%Y%m%d')
        filename = make_filename('AP_Inicial', nome, today)
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('ap_inicial.html',
                           farmacos=FARMACOS_BIOLOGICOS,
                           sinteticos=FARMACOS_SINTETICOS,
                           cids=CID_AP,
                           today=datetime.date.today().strftime('%d.%m.%Y'))


@app.route('/ea', methods=['GET', 'POST'])
def ea():
    if request.method == 'POST':
        data = request.form.to_dict()
        buf = generate_ea(data)
        nome = data.get('nome_paciente', 'paciente')
        tipo = data.get('tipo', 'Inicial')
        today = datetime.date.today().strftime('%Y%m%d')
        prefix = 'EA_Inicial' if tipo == 'Inicial' else 'EA_Manutencao'
        filename = make_filename(prefix, nome, today)
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('ea.html',
                           farmacos=FARMACOS_BIOLOGICOS_EA,
                           cids=CID_EA,
                           today=datetime.date.today().strftime('%d.%m.%Y'))


@app.route('/ea/inicial', methods=['GET', 'POST'])
def ea_inicial():
    if request.method == 'POST':
        data = request.form.to_dict()
        data['tipo'] = 'Inicial'
        buf = generate_ea(data)
        nome = data.get('nome_paciente', 'paciente')
        today = datetime.date.today().strftime('%Y%m%d')
        filename = make_filename('EA_Inicial', nome, today)
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('ea_inicial.html',
                           farmacos=FARMACOS_BIOLOGICOS_EA,
                           cids=CID_EA,
                           today=datetime.date.today().strftime('%d.%m.%Y'))


@app.route('/ea/manutencao', methods=['GET', 'POST'])
def ea_manutencao():
    if request.method == 'POST':
        data = request.form.to_dict()
        data['tipo'] = 'Manutencao'
        buf = generate_ea(data)
        nome = data.get('nome_paciente', 'paciente')
        today = datetime.date.today().strftime('%Y%m%d')
        filename = make_filename('EA_Manutencao', nome, today)
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('ea_manutencao.html',
                           farmacos=FARMACOS_BIOLOGICOS_EA,
                           cids=CID_EA,
                           today=datetime.date.today().strftime('%d.%m.%Y'))


@app.route('/ap/manutencao', methods=['GET', 'POST'])
def ap_manutencao():
    if request.method == 'POST':
        data = request.form.to_dict()
        buf = generate_ap_manutencao(data)
        nome = data.get('nome_paciente', 'paciente')
        today = datetime.date.today().strftime('%Y%m%d')
        filename = make_filename('AP_Manutencao', nome, today)
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return render_template('ap_manutencao.html',
                           farmacos=FARMACOS_BIOLOGICOS,
                           sinteticos=FARMACOS_SINTETICOS,
                           cids=CID_AP,
                           today=datetime.date.today().strftime('%d.%m.%Y'))


SYSTEM_PROMPT = """Você é o Dr. Fábio Batistella, médico reumatologista com CRM-RS 31746, \
atendendo em Passo Fundo – RS. Responda como se fosse o próprio Dr. Fábio, de forma \
acolhedora, clara e acessível para pacientes leigos.

Informações do consultório:
- Endereço: Edifício Vértice – Rua Capitão Araújo, 297, sala 808 – 8º andar, Passo Fundo – RS
- WhatsApp para agendamento: (54) 99959-7009
- Planos aceitos: Ipergs, Unimed, Capassemu, Cabergs e Particular
- Modalidades: atendimento presencial e teleconsulta
- Horário: segunda a sexta-feira, das 8h30 às 17h00

Especialidade:
- Reumatologista com foco em doenças autoimunes e inflamatórias
- Trata: Artrite Reumatoide, Artrite Psoriásica, Lúpus, Espondiloartrites, Gota, Osteoporose, \
Fibromialgia, Síndrome de Sjögren, Esclerodermia, entre outras doenças reumatológicas

Diretrizes:
- Responda dúvidas sobre reumatologia, sintomas, doenças e tratamentos de forma didática
- Para agendamentos ou dúvidas sobre consulta, oriente o paciente a entrar em contato pelo WhatsApp
- Não faça diagnósticos — oriente sempre que o diagnóstico precisa de consulta presencial
- Respostas curtas e objetivas (máximo 3 parágrafos)
- Sempre em português brasileiro
- Quando o paciente perguntar sobre agendamento, mencione o WhatsApp (54) 99959-7009"""


@app.route('/api/chat', methods=['POST'])
def chat():
    data = request.get_json()
    if not data or 'message' not in data:
        return jsonify({'error': 'Mensagem ausente'}), 400

    messages = data.get('history', [])
    messages.append({'role': 'user', 'content': data['message']})

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'API não configurada'}), 500

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=512,
            system=SYSTEM_PROMPT,
            messages=messages,
        )
        reply = response.content[0].text
        return jsonify({'reply': reply})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


TEXTO_ITEM3_AR = (
    'Já ter feito uso anterior dos medicamentos por pelo menos 3 a 6 meses: '
    'metotrexato ( ) ou hidroxicloroquina ( ) ou sulfassalazina ( ) ou leflunomida ( ), '
    'de forma isolada ou em associação.'
)
TEXTO_ITEM3_AP = (
    'Já ter feito uso anterior dos medicamentos: metotrexato ( ) ou sulfassalazina ( ) '
    'ou leflunomida ( ), de forma isolada ou em associação; '
    'pelo menos dois anti-inflamatórios não esteroidais ( ).'
)
TEXTO_ITEM3_EA = (
    'Já ter feito uso anterior de pelo menos dois anti-inflamatórios não esteroidais (AINEs) ( ) '
    'em doses plenas, por pelo menos 3 meses cada, sem resposta adequada ou com intolerância.'
)


@app.route('/tc/ar', methods=['GET', 'POST'])
def tc_ar():
    if request.method == 'POST':
        data = request.form.to_dict()
        return render_template('tc_print.html',
            doenca='Artrite Reumatoide',
            nome_paciente=data.get('nome_paciente', ''),
            idade=data.get('idade', ''),
            sexo=data.get('sexo', 'Feminino'),
            cid10=data.get('cid10', ''),
            data=data.get('data', datetime.date.today().strftime('%d/%m/%Y')),
            farmaco=data.get('farmaco', ''),
            farmacos_lista=FARMACOS_TC_AR,
            doctor=DOCTOR,
            texto_item3=TEXTO_ITEM3_AR,
        )
    return render_template('tc_ar.html',
        farmacos=FARMACOS_TC_AR,
        cids=CID_AR,
        today=datetime.date.today().strftime('%d/%m/%Y'),
    )


@app.route('/tc/ap', methods=['GET', 'POST'])
def tc_ap():
    if request.method == 'POST':
        data = request.form.to_dict()
        return render_template('tc_print.html',
            doenca='Artrite Psoriásica',
            nome_paciente=data.get('nome_paciente', ''),
            idade=data.get('idade', ''),
            sexo=data.get('sexo', 'Feminino'),
            cid10=data.get('cid10', ''),
            data=data.get('data', datetime.date.today().strftime('%d/%m/%Y')),
            farmaco=data.get('farmaco', ''),
            farmacos_lista=FARMACOS_TC_AP,
            doctor=DOCTOR,
            texto_item3=TEXTO_ITEM3_AP,
        )
    return render_template('tc_ap.html',
        farmacos=FARMACOS_TC_AP,
        cids=CID_AP,
        today=datetime.date.today().strftime('%d/%m/%Y'),
    )


@app.route('/tc/ea', methods=['GET', 'POST'])
def tc_ea():
    if request.method == 'POST':
        data = request.form.to_dict()
        return render_template('tc_print.html',
            doenca='Espondilite Anquilosante',
            nome_paciente=data.get('nome_paciente', ''),
            idade=data.get('idade', ''),
            sexo=data.get('sexo', 'Masculino'),
            cid10=data.get('cid10', ''),
            data=data.get('data', datetime.date.today().strftime('%d/%m/%Y')),
            farmaco=data.get('farmaco', ''),
            farmacos_lista=FARMACOS_TC_EA,
            doctor=DOCTOR,
            texto_item3=TEXTO_ITEM3_EA,
        )
    return render_template('tc_ea.html',
        farmacos=FARMACOS_TC_EA,
        cids=CID_EA,
        today=datetime.date.today().strftime('%d/%m/%Y'),
    )


SADT_PDF_ORIGINAL = "/Users/clinicabeneser/Downloads/EXAMES UNIMED.pdf"
SADT_ORIG_W, SADT_ORIG_H = 2058.0, 2924.0
SADT_LAND_W, SADT_LAND_H = SADT_ORIG_H, SADT_ORIG_W
SADT_CAMPOS = {
    "nome":        {"x": 1217, "y": 1721, "tamanho": 34},
    "ind_clinica": {"x": 1375, "y": 1341, "tamanho": 34},
    "descricao":   {"x":  787, "y": 1274, "tamanho": 34},
}


@app.route('/unimed/sadt', methods=['GET', 'POST'])
def unimed_sadt():
    if request.method == 'POST':
        nome      = request.form.get('nome', '').strip()
        ind       = request.form.get('ind_clinica', '').strip()
        desc      = request.form.get('descricao', '').strip()
        try:
            reader = PdfReader(SADT_PDF_ORIGINAL)
            page   = reader.pages[0]
            t = Transformation((0, 1, -1, 0, SADT_ORIG_H, 0))
            page.add_transformation(t)
            page.mediabox = RectangleObject([0, 0, SADT_LAND_W, SADT_LAND_H])

            buf = io.BytesIO()
            c = canvas.Canvas(buf, pagesize=(SADT_LAND_W, SADT_LAND_H))
            c.setFillColor(black)
            for chave, texto in [("nome", nome), ("ind_clinica", ind), ("descricao", desc)]:
                if texto:
                    cfg = SADT_CAMPOS[chave]
                    c.setFont("Helvetica", cfg["tamanho"])
                    c.drawString(cfg["x"], cfg["y"], texto)
            c.save()

            buf.seek(0)
            overlay = PdfReader(buf)
            page.merge_page(overlay.pages[0])

            writer = PdfWriter()
            writer.add_page(page)
            out = io.BytesIO()
            writer.write(out)
            out.seek(0)
            return send_file(out, mimetype="application/pdf",
                             download_name="GUIA_PREENCHIDA.pdf",
                             as_attachment=False)
        except Exception as e:
            return render_template('unimed_sadt.html',
                                   nome=nome, ind_clinica=ind,
                                   descricao=desc, erro=str(e))
    return render_template('unimed_sadt.html',
                           nome='', ind_clinica='', descricao='', erro=None)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
