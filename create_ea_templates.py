"""Script para criar os templates Excel de EA (Espondilite Anquilosante)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, 'excel_templates')

# Estilos
HEADER_FILL  = PatternFill("solid", fgColor="C00000")
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
SUB_FILL     = PatternFill("solid", fgColor="D9E1F2")
LABEL_FILL   = PatternFill("solid", fgColor="F2F2F2")
WHITE_FONT   = Font(color="FFFFFF", bold=True, size=10)
BOLD_FONT    = Font(bold=True, size=10)
NORMAL_FONT  = Font(size=10)
THIN_BORDER  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, coord, value, font=None, fill=None, align=None, border=None):
    ws[coord] = value
    if font:   ws[coord].font = font
    if fill:   ws[coord].fill = fill
    if align:  ws[coord].alignment = align
    if border: ws[coord].border = border

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def create_ea_inicial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EA Inicial"

    set_col_widths(ws, {'A': 4, 'B': 20, 'C': 28, 'D': 14, 'E': 22, 'F': 18})

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Título
    ws.merge_cells('A1:F1')
    set_cell(ws, 'A1',
             'IPERGS — SOLICITAÇÃO DE MEDICAMENTO IMUNOBIOLÓGICO\nEspondilite Anquilosante — SOLICITAÇÃO INICIAL',
             font=Font(color="FFFFFF", bold=True, size=12),
             fill=HEADER_FILL, align=center)
    ws.row_dimensions[1].height = 36

    # I — Médico
    ws.merge_cells('A2:F2')
    set_cell(ws, 'A2', 'I — IDENTIFICAÇÃO DO MÉDICO SOLICITANTE',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    for row, (label, col_val) in enumerate([
        ('Médico:', 'C4'), ('CRM-RS:', 'C5'), ('Especialidade:', 'E5'), ('Telefone:', 'C6')
    ], start=4):
        pass

    ws['B4'] = 'Médico:';       ws['B4'].font = BOLD_FONT; ws['B4'].fill = LABEL_FILL
    ws['B5'] = 'CRM-RS:';       ws['B5'].font = BOLD_FONT; ws['B5'].fill = LABEL_FILL
    ws['D5'] = 'Especialidade:';ws['D5'].font = BOLD_FONT; ws['D5'].fill = LABEL_FILL
    ws['B6'] = 'Telefone:';     ws['B6'].font = BOLD_FONT; ws['B6'].fill = LABEL_FILL
    for coord in ['C4','C5','E5','C6']:
        ws[coord].border = THIN_BORDER
        ws[coord].font   = NORMAL_FONT

    # II — Paciente
    ws.merge_cells('A7:F7')
    set_cell(ws, 'A7', 'II — IDENTIFICAÇÃO DO PACIENTE',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B8']  = 'Nome:';     ws['B8'].font  = BOLD_FONT; ws['B8'].fill  = LABEL_FILL
    ws['B9']  = 'Idade:';    ws['B9'].font  = BOLD_FONT; ws['B9'].fill  = LABEL_FILL
    ws['D9']  = 'Sexo:';     ws['D9'].font  = BOLD_FONT; ws['D9'].fill  = LABEL_FILL
    ws['B10'] = 'Telefone:'; ws['B10'].font = BOLD_FONT; ws['B10'].fill = LABEL_FILL
    for coord in ['C8','C9','E9','C10']:
        ws[coord].border = THIN_BORDER
        ws[coord].font   = NORMAL_FONT

    # III — Histórico
    ws.merge_cells('A11:F11')
    set_cell(ws, 'A11', 'III — HISTÓRICO DA DOENÇA',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B12'] = 'Data diagnóstico:'; ws['B12'].font = BOLD_FONT; ws['B12'].fill = LABEL_FILL
    ws['B13'] = 'CID-10:';           ws['B13'].font = BOLD_FONT; ws['B13'].fill = LABEL_FILL
    ws['B14'] = 'HLA-B27:';          ws['B14'].font = BOLD_FONT; ws['B14'].fill = LABEL_FILL
    ws['B15'] = 'Forma da doença:';  ws['B15'].font = BOLD_FONT; ws['B15'].fill = LABEL_FILL
    ws['B16'] = 'Critério diagnóst.';ws['B16'].font = BOLD_FONT; ws['B16'].fill = LABEL_FILL
    for coord in ['C12','C13','C14','C15','C16']:
        ws[coord].border = THIN_BORDER
        ws[coord].font   = NORMAL_FONT

    # IV — Imagem
    ws.merge_cells('A17:F17')
    set_cell(ws, 'A17', 'IV — EXAMES DE IMAGEM',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    img_labels = [
        (18, 'RX Sacroilíacas:', 'D18', 'Grau sacroileíte:'),
        (19, 'RMN Sacroilíacas:', 'D19', 'Edema ósseo (STIR):'),
        (20, '', 'D20', 'Erosões/Esclerose:'),
        (21, 'RX Coluna:', 'D21', 'Sindesmófitos:'),
        (22, '', 'D22', 'Coluna em bambu:'),
    ]
    for row, label_b, coord_d, label_d in img_labels:
        if label_b:
            ws[f'B{row}'] = label_b; ws[f'B{row}'].font = BOLD_FONT; ws[f'B{row}'].fill = LABEL_FILL
        ws[f'C{row}'].border = THIN_BORDER; ws[f'C{row}'].font = NORMAL_FONT
        ws[f'{coord_d}'] = label_d; ws[f'{coord_d}'].font = BOLD_FONT; ws[f'{coord_d}'].fill = LABEL_FILL
        ws[f'E{row}'].border = THIN_BORDER; ws[f'E{row}'].font = NORMAL_FONT

    # V — Atividade
    ws.merge_cells('A23:F23')
    set_cell(ws, 'A23', 'V — ATIVIDADE DA DOENÇA',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws.merge_cells('A24:F24')
    set_cell(ws, 'A24', 'BASDAI — Bath Ankylosing Spondylitis Disease Activity Index',
             font=BOLD_FONT, fill=SUB_FILL, align=left)

    basdai_q = [
        'Q1. Nível geral de fadiga/cansaço',
        'Q2. Dor no pescoço, costas ou quadris',
        'Q3. Dor/inchaço em outras articulações',
        'Q4. Desconforto em áreas sensíveis ao toque',
        'Q5. Nível de rigidez matinal ao acordar',
        'Q6. Duração da rigidez matinal (0=nenhuma; 10=≥2h)',
    ]
    for i, q in enumerate(basdai_q):
        row = 25 + i
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = q
        ws[f'B{row}'].font = NORMAL_FONT
        ws[f'E{row}'] = 'Valor (0–10):'
        ws[f'E{row}'].font = BOLD_FONT
        ws[f'E{row}'].fill = LABEL_FILL
        ws[f'F{row}'].border = THIN_BORDER

    ws['B31'] = 'BASDAI (calculado):'
    ws['B31'].font = BOLD_FONT; ws['B31'].fill = SUB_FILL
    ws['C31'].border = THIN_BORDER; ws['C31'].font = BOLD_FONT

    ws.merge_cells('A32:F32')
    set_cell(ws, 'A32', 'ASDAS — Ankylosing Spondylitis Disease Activity Score',
             font=BOLD_FONT, fill=SUB_FILL, align=left)

    ws['B33'] = 'Aval. global paciente (PGA 0–10):'; ws['B33'].font = BOLD_FONT; ws['B33'].fill = LABEL_FILL
    ws['C33'].border = THIN_BORDER
    ws['B34'] = 'PCR (mg/dL):'; ws['B34'].font = BOLD_FONT; ws['B34'].fill = LABEL_FILL
    ws['C34'].border = THIN_BORDER
    ws['B35'] = 'VSG/VHS (mm/h):'; ws['B35'].font = BOLD_FONT; ws['B35'].fill = LABEL_FILL
    ws['C35'].border = THIN_BORDER
    ws['B36'] = 'Índice utilizado:'; ws['B36'].font = BOLD_FONT; ws['B36'].fill = LABEL_FILL
    ws['C36'].border = THIN_BORDER
    ws['B37'] = 'Valor do índice:'; ws['B37'].font = BOLD_FONT; ws['B37'].fill = LABEL_FILL
    ws['C37'].border = THIN_BORDER; ws['C37'].font = BOLD_FONT

    # VI — Tratamento anterior
    ws.merge_cells('A38:F38')
    set_cell(ws, 'A38', 'VI — HISTÓRICO DO TRATAMENTO (AINEs e DMARDs anteriores)',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws.merge_cells('A39:F39')
    set_cell(ws, 'A39', 'Exige falha de pelo menos 2 AINEs em doses adequadas por no mínimo 3 meses cada.',
             font=Font(italic=True, size=9), align=left)

    for i in range(4):
        row = 40 + i
        ws[f'A{row}'] = str(i + 1);   ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'] = 'Fármaco:';   ws[f'B{row}'].font = BOLD_FONT; ws[f'B{row}'].fill = LABEL_FILL
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'D{row}'] = 'Posologia:'; ws[f'D{row}'].font = BOLD_FONT; ws[f'D{row}'].fill = LABEL_FILL
        ws[f'E{row}'] = 'Período:';   ws[f'E{row}'].font = BOLD_FONT; ws[f'E{row}'].fill = LABEL_FILL
        ws[f'F{row}'].border = THIN_BORDER

    # VII — Terapia proposta
    ws.merge_cells('A44:F44')
    set_cell(ws, 'A44', 'VII — TERAPIA IMUNOBIOLÓGICA PROPOSTA',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B45'] = 'Fármaco:';     ws['B45'].font = BOLD_FONT; ws['B45'].fill = LABEL_FILL
    ws['C45'].border = THIN_BORDER
    ws['B46'] = 'Posologia:';   ws['B46'].font = BOLD_FONT; ws['B46'].fill = LABEL_FILL
    ws['C46'].border = THIN_BORDER
    ws['B47'] = 'Peso:';        ws['B47'].font = BOLD_FONT; ws['B47'].fill = LABEL_FILL
    ws['C47'].border = THIN_BORDER
    ws['B48'] = 'RX tórax/PPD:';ws['B48'].font = BOLD_FONT; ws['B48'].fill = LABEL_FILL
    ws['C48'].border = THIN_BORDER

    # VIII — Observações
    ws.merge_cells('A49:F49')
    set_cell(ws, 'A49', 'VIII — OBSERVAÇÕES E DATA',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws.merge_cells('B50:F52')
    ws['B50'].border = THIN_BORDER
    ws['B50'].alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[50].height = 45

    ws['B54'] = 'Data:'; ws['B54'].font = BOLD_FONT; ws['B54'].fill = LABEL_FILL
    ws['C54'].border = THIN_BORDER

    ws.merge_cells('B56:F57')
    set_cell(ws, 'B56',
             '________________________________________\n' + 'Dr. Fábio Batistella — CRM-RS 31746\nReumatologista',
             font=NORMAL_FONT,
             align=Alignment(horizontal='center', vertical='center', wrap_text=True))
    ws.row_dimensions[56].height = 48

    os.makedirs(EXCEL_DIR, exist_ok=True)
    path = os.path.join(EXCEL_DIR, 'ea_inicial.xlsx')
    wb.save(path)
    print(f"Criado: {path}")


def create_ea_manutencao():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EA Manutencao"

    set_col_widths(ws, {'A': 4, 'B': 22, 'C': 28, 'D': 14, 'E': 22, 'F': 18})

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Título
    ws.merge_cells('A1:F1')
    set_cell(ws, 'A1',
             'IPERGS — SOLICITAÇÃO DE MEDICAMENTO IMUNOBIOLÓGICO\nEspondilite Anquilosante — MANUTENÇÃO',
             font=Font(color="FFFFFF", bold=True, size=12),
             fill=HEADER_FILL, align=center)
    ws.row_dimensions[1].height = 36

    # I — Médico
    ws.merge_cells('A2:F2')
    set_cell(ws, 'A2', 'I — IDENTIFICAÇÃO DO MÉDICO SOLICITANTE',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B4'] = 'Médico:';        ws['B4'].font = BOLD_FONT; ws['B4'].fill = LABEL_FILL
    ws['B5'] = 'CRM-RS:';        ws['B5'].font = BOLD_FONT; ws['B5'].fill = LABEL_FILL
    ws['D5'] = 'Especialidade:'; ws['D5'].font = BOLD_FONT; ws['D5'].fill = LABEL_FILL
    ws['B6'] = 'Telefone:';      ws['B6'].font = BOLD_FONT; ws['B6'].fill = LABEL_FILL
    for coord in ['C4','C5','E5','C6']:
        ws[coord].border = THIN_BORDER
        ws[coord].font   = NORMAL_FONT

    # II — Paciente
    ws.merge_cells('A7:F7')
    set_cell(ws, 'A7', 'II — DADOS DO PACIENTE',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B8']  = 'Nome:';                ws['B8'].font  = BOLD_FONT; ws['B8'].fill  = LABEL_FILL
    ws['B9']  = 'Idade:';               ws['B9'].font  = BOLD_FONT; ws['B9'].fill  = LABEL_FILL
    ws['D9']  = 'Sexo:';                ws['D9'].font  = BOLD_FONT; ws['D9'].fill  = LABEL_FILL
    ws['B10'] = 'Início tratamento:';   ws['B10'].font = BOLD_FONT; ws['B10'].fill = LABEL_FILL
    ws['B11'] = 'CID-10:';              ws['B11'].font = BOLD_FONT; ws['B11'].fill = LABEL_FILL
    ws['B12'] = 'Peso:';                ws['B12'].font = BOLD_FONT; ws['B12'].fill = LABEL_FILL
    for coord in ['C8','C9','E9','C10','C11','C12']:
        ws[coord].border = THIN_BORDER
        ws[coord].font   = NORMAL_FONT

    # III — Atividade
    ws.merge_cells('A13:F13')
    set_cell(ws, 'A13', 'III — ATIVIDADE DA DOENÇA',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B14'] = 'Índice utilizado:'; ws['B14'].font = BOLD_FONT; ws['B14'].fill = LABEL_FILL
    ws['C14'].border = THIN_BORDER

    ws.merge_cells('A15:F15')
    set_cell(ws, 'A15', 'BASDAI — Questões (EVA 0–10)',
             font=BOLD_FONT, fill=SUB_FILL, align=left)

    basdai_q = [
        'Q1. Fadiga/cansaço geral', 'Q2. Dor no pescoço, costas ou quadris',
        'Q3. Dor/inchaço em outras articulações', 'Q4. Desconforto em áreas sensíveis',
        'Q5. Rigidez matinal (nível)', 'Q6. Rigidez matinal (duração)',
    ]
    for i, q in enumerate(basdai_q):
        row = 16 + i
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = q; ws[f'B{row}'].font = NORMAL_FONT
        ws[f'E{row}'] = 'Valor:'; ws[f'E{row}'].font = BOLD_FONT; ws[f'E{row}'].fill = LABEL_FILL
        ws[f'F{row}'].border = THIN_BORDER

    ws['B22'] = 'BASDAI (calculado):'; ws['B22'].font = BOLD_FONT; ws['B22'].fill = SUB_FILL
    ws['C22'].border = THIN_BORDER; ws['C22'].font = BOLD_FONT

    ws['B23'] = 'Aval. global paciente (PGA 0–10):'; ws['B23'].font = BOLD_FONT; ws['B23'].fill = LABEL_FILL
    ws['C23'].border = THIN_BORDER
    ws['B24'] = 'PCR (mg/dL):'; ws['B24'].font = BOLD_FONT; ws['B24'].fill = LABEL_FILL
    ws['C24'].border = THIN_BORDER
    ws['B25'] = 'VSG/VHS (mm/h):'; ws['B25'].font = BOLD_FONT; ws['B25'].fill = LABEL_FILL
    ws['C25'].border = THIN_BORDER
    ws['B26'] = 'Valor do índice:'; ws['B26'].font = BOLD_FONT; ws['B26'].fill = LABEL_FILL
    ws['C26'].border = THIN_BORDER; ws['C26'].font = BOLD_FONT

    # IV — Evolução
    ws.merge_cells('A27:F27')
    set_cell(ws, 'A27', 'IV — EVOLUÇÃO CLÍNICA',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B28'] = 'Boa resposta terapêutica?'; ws['B28'].font = BOLD_FONT; ws['B28'].fill = LABEL_FILL
    ws['C28'].border = THIN_BORDER
    ws.merge_cells('B29:F30')
    ws['B29'] = 'Falha terapêutica / evento adverso:'
    ws['B29'].font = BOLD_FONT; ws['B29'].fill = LABEL_FILL
    ws.merge_cells('B30:F30')
    ws['B30'].border = THIN_BORDER
    ws['B30'].alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[30].height = 30

    # V — Recomendação
    ws.merge_cells('A31:F31')
    set_cell(ws, 'A31', 'V — RECOMENDAÇÃO',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws['B32'] = 'Manter tratamento?';   ws['B32'].font = BOLD_FONT; ws['B32'].fill = LABEL_FILL
    ws['C32'].border = THIN_BORDER
    ws['D32'] = 'Modificar tratamento?';ws['D32'].font = BOLD_FONT; ws['D32'].fill = LABEL_FILL
    ws['E32'].border = THIN_BORDER
    ws['B33'] = 'Fármaco:';             ws['B33'].font = BOLD_FONT; ws['B33'].fill = LABEL_FILL
    ws['C33'].border = THIN_BORDER
    ws['B34'] = 'Posologia:';           ws['B34'].font = BOLD_FONT; ws['B34'].fill = LABEL_FILL
    ws['C34'].border = THIN_BORDER

    # VI — Observações
    ws.merge_cells('A35:F35')
    set_cell(ws, 'A35', 'VI — OBSERVAÇÕES E DATA',
             font=WHITE_FONT, fill=SECTION_FILL, align=center)

    ws.merge_cells('B36:F38')
    ws['B36'].border = THIN_BORDER
    ws['B36'].alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[36].height = 45

    ws['B40'] = 'Data:'; ws['B40'].font = BOLD_FONT; ws['B40'].fill = LABEL_FILL
    ws['C40'].border = THIN_BORDER

    ws.merge_cells('B42:F43')
    set_cell(ws, 'B42',
             '________________________________________\n' + 'Dr. Fábio Batistella — CRM-RS 31746\nReumatologista',
             font=NORMAL_FONT,
             align=Alignment(horizontal='center', vertical='center', wrap_text=True))
    ws.row_dimensions[42].height = 48

    path = os.path.join(EXCEL_DIR, 'ea_manutencao.xlsx')
    wb.save(path)
    print(f"Criado: {path}")


if __name__ == '__main__':
    create_ea_inicial()
    create_ea_manutencao()
    print("Templates EA criados com sucesso.")
